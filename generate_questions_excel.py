#!/usr/bin/env python3

import openpyxl
from openpyxl.utils import get_column_letter

# 创建一个新的工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Questions"

# 设置表头
headers = ["level", "subject", "type", "content", "options", "answer", "explanation"]
for col, header in enumerate(headers, 1):
    worksheet[get_column_letter(col) + "1"] = header

# 示例题目数据
questions = [
    # 入门级 - 算法基础 (10道)
    ["entry", "1", "single", "以下哪个是Python中的内置数据类型？", '{"A":"Array","B":"List","C":"HashTable","D":"Map"}', "B", "Python中的内置数据类型包括List，而Array、HashTable、Map都不是内置类型"],
    ["entry", "1", "judge", "Python中的列表是可变的数据类型。", '{"A":"正确","B":"错误"}', "A", "Python中的列表(list)是可变的数据类型，可以通过索引修改元素"],
    ["entry", "1", "single", "以下哪个不是Python中的循环结构？", '{"A":"for","B":"while","C":"do-while","D":"都不是"}', "C", "Python中没有do-while循环结构"],
    ["entry", "1", "multiple", "以下哪些是Python中的内置函数？", '{"A":"print()","B":"len()","C":"range()","D":"sort()"}', "ABC", "print()、len()、range()都是Python中的内置函数，sort()是列表的方法"],
    ["entry", "1", "single", "Python中，以下哪个符号用于单行注释？", '{"A":"/*","B":"//","C":"#","D":"--"}', "C", "Python中使用#符号进行单行注释"],
    ["entry", "1", "judge", "Python是一种解释型语言。", '{"A":"正确","B":"错误"}', "A", "Python是一种解释型语言，不需要编译成机器码"],
    ["entry", "1", "single", "以下哪个是Python中的条件判断语句？", '{"A":"if-else","B":"switch-case","C":"select-case","D":"以上都不是"}', "A", "Python中使用if-else进行条件判断，没有switch-case语句"],
    ["entry", "1", "multiple", "以下哪些是Python中的数据结构？", '{"A":"列表","B":"元组","C":"字典","D":"集合"}', "ABCD", "Python中的数据结构包括列表、元组、字典和集合"],
    ["entry", "1", "single", "Python中，以下哪个方法用于将字符串转换为整数？", '{"A":"str()","B":"int()","C":"float()","D":"bool()"}', "B", "Python中使用int()函数将字符串转换为整数"],
    ["entry", "1", "judge", "Python中的变量不需要声明类型。", '{"A":"正确","B":"错误"}', "A", "Python是动态类型语言，变量不需要声明类型"],
    
    # 入门级 - 编程语言 (10道)
    ["entry", "2", "single", "以下哪个是JavaScript中的闭包？", '{"A":"一个函数返回另一个函数","B":"一个函数调用自身","C":"一个函数接收另一个函数作为参数","D":"一个函数没有参数"}', "A", "闭包是指有权访问另一个函数作用域中变量的函数，通常是一个函数返回另一个函数"],
    ["entry", "2", "multiple", "以下哪些是JavaScript中的基本数据类型？", '{"A":"String","B":"Number","C":"Object","D":"Boolean"}', "ABD", "JavaScript中的基本数据类型包括String、Number、Boolean、Null、Undefined、Symbol、BigInt，Object是引用类型"],
    ["entry", "2", "single", "JavaScript中，以下哪个符号用于声明变量？", '{"A":"var","B":"let","C":"const","D":"以上都是"}', "D", "JavaScript中可以使用var、let、const声明变量"],
    ["entry", "2", "judge", "JavaScript是一种面向对象的编程语言。", '{"A":"正确","B":"错误"}', "A", "JavaScript是一种基于原型的面向对象编程语言"],
    ["entry", "2", "single", "JavaScript中，以下哪个方法用于添加事件监听器？", '{"A":"addEventListener()","B":"on()","C":"bind()","D":"attachEvent()"}', "A", "现代JavaScript中使用addEventListener()添加事件监听器"],
    ["entry", "2", "multiple", "以下哪些是JavaScript中的循环结构？", '{"A":"for","B":"while","C":"do-while","D":"for-in"}', "ABCD", "JavaScript中的循环结构包括for、while、do-while和for-in"],
    ["entry", "2", "single", "JavaScript中，以下哪个方法用于创建数组？", '{"A":"new Array()","B":"[]","C":"Array()","D":"以上都是"}', "D", "JavaScript中可以使用new Array()、[]或Array()创建数组"],
    ["entry", "2", "judge", "JavaScript中的对象是键值对的集合。", '{"A":"正确","B":"错误"}', "A", "JavaScript中的对象是键值对的集合"],
    ["entry", "2", "single", "JavaScript中，以下哪个方法用于字符串拼接？", '{"A":"+","B":"concat()","C":"join()","D":"以上都是"}', "D", "JavaScript中可以使用+运算符、concat()方法或join()方法进行字符串拼接"],
    ["entry", "2", "multiple", "以下哪些是JavaScript中的DOM操作方法？", '{"A":"getElementById()","B":"querySelector()","C":"createElement()","D":"appendChild()"}', "ABCD", "这些都是JavaScript中的DOM操作方法"],
    
    # 入门级 - 标准规范 (10道)
    ["entry", "3", "single", "在HTTP协议中，以下哪个状态码表示请求成功？", '{"A":"404","B":"500","C":"200","D":"403"}', "C", "HTTP 200状态码表示请求成功，404表示资源未找到，500表示服务器内部错误，403表示禁止访问"],
    ["entry", "3", "judge", "RESTful API中，GET请求可以修改服务器端资源。", '{"A":"正确","B":"错误"}', "B", "RESTful API中，GET请求应该是幂等的，不应该修改服务器端资源"],
    ["entry", "3", "single", "以下哪个是HTTP协议中的请求方法？", '{"A":"GET","B":"POST","C":"PUT","D":"以上都是"}', "D", "GET、POST、PUT都是HTTP协议中的请求方法"],
    ["entry", "3", "multiple", "以下哪些是HTTP协议中的常见头部字段？", '{"A":"Content-Type","B":"Authorization","C":"Cache-Control","D":"User-Agent"}', "ABCD", "这些都是HTTP协议中的常见头部字段"],
    ["entry", "3", "single", "在RESTful API中，以下哪个方法用于创建资源？", '{"A":"GET","B":"POST","C":"PUT","D":"DELETE"}', "B", "RESTful API中，POST方法用于创建资源"],
    ["entry", "3", "judge", "HTTPS是HTTP的安全版本，使用SSL/TLS加密。", '{"A":"正确","B":"错误"}', "A", "HTTPS是HTTP的安全版本，使用SSL/TLS进行加密"],
    ["entry", "3", "single", "以下哪个是HTTP协议的默认端口？", '{"A":"80","B":"443","C":"8080","D":"3306"}', "A", "HTTP协议的默认端口是80，HTTPS的默认端口是443"],
    ["entry", "3", "multiple", "以下哪些是RESTful API的设计原则？", '{"A":"资源标识","B":"统一接口","C":"无状态","D":"缓存"}', "ABCD", "这些都是RESTful API的设计原则"],
    ["entry", "3", "single", "在HTTP协议中，以下哪个状态码表示资源未找到？", '{"A":"200","B":"404","C":"500","D":"403"}', "B", "HTTP 404状态码表示资源未找到"],
    ["entry", "3", "judge", "RESTful API中，URL应该包含动词。", '{"A":"正确","B":"错误"}', "B", "RESTful API中，URL应该包含名词（资源），而不是动词"],
    
    # 入门级 - 设计模式 (10道)
    ["entry", "4", "single", "以下哪个设计模式属于创建型模式？", '{"A":"单例模式","B":"适配器模式","C":"观察者模式","D":"策略模式"}', "A", "单例模式属于创建型模式，适配器模式属于结构型模式，观察者模式和策略模式属于行为型模式"],
    ["entry", "4", "multiple", "以下哪些是SOLID原则的内容？", '{"A":"单一职责原则","B":"开放封闭原则","C":"里氏替换原则","D":"依赖倒置原则"}', "ABCD", "SOLID原则包括：单一职责原则(S)、开放封闭原则(O)、里氏替换原则(L)、接口隔离原则(I)、依赖倒置原则(D)"],
    ["entry", "4", "single", "以下哪个设计模式属于结构型模式？", '{"A":"工厂模式","B":"适配器模式","C":"观察者模式","D":"命令模式"}', "B", "适配器模式属于结构型模式，工厂模式属于创建型模式，观察者模式和命令模式属于行为型模式"],
    ["entry", "4", "judge", "单例模式确保一个类只有一个实例。", '{"A":"正确","B":"错误"}', "A", "单例模式的目的是确保一个类只有一个实例"],
    ["entry", "4", "single", "以下哪个设计模式属于行为型模式？", '{"A":"单例模式","B":"装饰器模式","C":"观察者模式","D":"工厂模式"}', "C", "观察者模式属于行为型模式，单例模式和工厂模式属于创建型模式，装饰器模式属于结构型模式"],
    ["entry", "4", "multiple", "以下哪些是创建型设计模式？", '{"A":"单例模式","B":"工厂模式","C":"抽象工厂模式","D":"建造者模式"}', "ABCD", "这些都属于创建型设计模式"],
    ["entry", "4", "single", "以下哪个设计模式用于将一个类的接口转换成客户端所期望的另一个接口？", '{"A":"适配器模式","B":"装饰器模式","C":"观察者模式","D":"策略模式"}', "A", "适配器模式用于将一个类的接口转换成客户端所期望的另一个接口"],
    ["entry", "4", "judge", "装饰器模式允许在不修改原有对象结构的情况下，动态地给对象添加额外的责任。", '{"A":"正确","B":"错误"}', "A", "装饰器模式的核心思想就是在不修改原有对象结构的情况下，动态地给对象添加额外的责任"],
    ["entry", "4", "single", "以下哪个设计模式用于定义对象间的一种一对多的依赖关系，当一个对象状态发生变化时，所有依赖于它的对象都得到通知并被自动更新？", '{"A":"适配器模式","B":"装饰器模式","C":"观察者模式","D":"策略模式"}', "C", "观察者模式用于定义对象间的一种一对多的依赖关系"],
    ["entry", "4", "multiple", "以下哪些是行为型设计模式？", '{"A":"观察者模式","B":"策略模式","C":"命令模式","D":"迭代器模式"}', "ABCD", "这些都属于行为型设计模式"],
    
    # 工作级 - 算法基础 (10道)
    ["work", "1", "single", "以下哪个排序算法的时间复杂度在最坏情况下是O(n^2)？", '{"A":"快速排序","B":"归并排序","C":"堆排序","D":"基数排序"}', "A", "快速排序在最坏情况下的时间复杂度是O(n^2)，归并排序和堆排序是O(nlogn)，基数排序是O(nk)"],
    ["work", "1", "multiple", "以下哪些算法可以用于查找最短路径？", '{"A":"Dijkstra算法","B":"Floyd-Warshall算法","C":"Prim算法","D":"Kruskal算法"}', "AB", "Dijkstra算法和Floyd-Warshall算法可以用于查找最短路径，Prim算法和Kruskal算法用于最小生成树"],
    ["work", "1", "single", "以下哪个算法用于在有序数组中查找元素？", '{"A":"线性查找","B":"二分查找","C":"哈希查找","D":"以上都不是"}', "B", "二分查找用于在有序数组中查找元素，时间复杂度为O(logn)"],
    ["work", "1", "judge", "二分查找的时间复杂度是O(logn)。", '{"A":"正确","B":"错误"}', "A", "二分查找的时间复杂度是O(logn)"],
    ["work", "1", "single", "以下哪个是分治算法的例子？", '{"A":"快速排序","B":"冒泡排序","C":"插入排序","D":"选择排序"}', "A", "快速排序是分治算法的例子，将大问题分解为小问题解决"],
    ["work", "1", "multiple", "以下哪些是动态规划的应用场景？", '{"A":"背包问题","B":"最长公共子序列","C":"最短路径","D":"斐波那契数列"}', "ABCD", "这些都是动态规划的应用场景"],
    ["work", "1", "single", "以下哪个算法用于生成最小生成树？", '{"A":"Dijkstra算法","B":"Floyd-Warshall算法","C":"Prim算法","D":"Kruskal算法"}', "C", "Prim算法和Kruskal算法都用于生成最小生成树"],
    ["work", "1", "judge", "快速排序是稳定的排序算法。", '{"A":"正确","B":"错误"}', "B", "快速排序是不稳定的排序算法"],
    ["work", "1", "single", "以下哪个数据结构用于实现先进先出（FIFO）？", '{"A":"栈","B":"队列","C":"链表","D":"树"}', "B", "队列用于实现先进先出（FIFO）"],
    ["work", "1", "multiple", "以下哪些数据结构可以用于实现栈？", '{"A":"数组","B":"链表","C":"队列","D":"树"}', "AB", "栈可以用数组或链表实现"],
    
    # 工作级 - 编程语言 (10道)
    ["work", "2", "single", "在Java中，以下哪个关键字用于创建对象？", '{"A":"new","B":"create","C":"make","D":"instance"}', "A", "在Java中，使用new关键字创建对象"],
    ["work", "2", "judge", "Java中的String类是可变的。", '{"A":"正确","B":"错误"}', "B", "Java中的String类是不可变的，每次修改都会创建新的String对象"],
    ["work", "2", "single", "Java中，以下哪个关键字用于继承？", '{"A":"implements","B":"extends","C":"inherits","D":"以上都不是"}', "B", "Java中使用extends关键字进行继承"],
    ["work", "2", "multiple", "以下哪些是Java中的访问修饰符？", '{"A":"public","B":"private","C":"protected","D":"default"}', "ABCD", "这些都是Java中的访问修饰符"],
    ["work", "2", "single", "Java中，以下哪个接口用于实现集合的迭代？", '{"A":"Collection","B":"List","C":"Set","D":"Iterator"}', "D", "Iterator接口用于实现集合的迭代"],
    ["work", "2", "judge", "Java是一种编译型语言。", '{"A":"正确","B":"错误"}', "A", "Java是一种编译型语言，需要编译成字节码"],
    ["work", "2", "single", "Java中，以下哪个方法用于启动线程？", '{"A":"run()","B":"start()","C":"begin()","D":"execute()"}', "B", "Java中使用start()方法启动线程"],
    ["work", "2", "multiple", "以下哪些是Java中的异常处理关键字？", '{"A":"try","B":"catch","C":"finally","D":"throw"}', "ABCD", "这些都是Java中的异常处理关键字"],
    ["work", "2", "single", "Java中，以下哪个类是所有类的父类？", '{"A":"Object","B":"Class","C":"System","D":"Runtime"}', "A", "Object类是Java中所有类的父类"],
    ["work", "2", "judge", "Java中的接口可以有方法实现。", '{"A":"正确","B":"错误"}', "A", "Java 8及以上版本中，接口可以有默认方法和静态方法的实现"],
    
    # 工作级 - 标准规范 (10道)
    ["work", "3", "single", "以下哪个是SQL中的聚合函数？", '{"A":"SELECT","B":"FROM","C":"SUM","D":"WHERE"}', "C", "SUM是SQL中的聚合函数，用于计算总和，其他选项都是SQL语句的关键字"],
    ["work", "3", "multiple", "以下哪些是SQL中的约束？", '{"A":"PRIMARY KEY","B":"FOREIGN KEY","C":"UNIQUE","D":"NOT NULL"}', "ABCD", "这些都是SQL中的约束，用于保证数据的完整性和一致性"],
    ["work", "3", "single", "SQL中，以下哪个语句用于查询数据？", '{"A":"INSERT","B":"UPDATE","C":"DELETE","D":"SELECT"}', "D", "SELECT语句用于查询数据"],
    ["work", "3", "judge", "SQL中的JOIN用于连接两个或多个表。", '{"A":"正确","B":"错误"}', "A", "SQL中的JOIN用于连接两个或多个表"],
    ["work", "3", "single", "SQL中，以下哪个子句用于对结果集进行排序？", '{"A":"WHERE","B":"GROUP BY","C":"ORDER BY","D":"HAVING"}', "C", "ORDER BY子句用于对结果集进行排序"],
    ["work", "3", "multiple", "以下哪些是SQL中的JOIN类型？", '{"A":"INNER JOIN","B":"LEFT JOIN","C":"RIGHT JOIN","D":"FULL JOIN"}', "ABCD", "这些都是SQL中的JOIN类型"],
    ["work", "3", "single", "SQL中，以下哪个语句用于创建表？", '{"A":"CREATE TABLE","B":"ALTER TABLE","C":"DROP TABLE","D":"TRUNCATE TABLE"}', "A", "CREATE TABLE语句用于创建表"],
    ["work", "3", "judge", "SQL中的事务是原子的、一致的、隔离的和持久的（ACID）。", '{"A":"正确","B":"错误"}', "A", "SQL中的事务具有ACID特性"],
    ["work", "3", "single", "SQL中，以下哪个函数用于计算平均值？", '{"A":"SUM()","B":"AVG()","C":"COUNT()","D":"MAX()"}', "B", "AVG()函数用于计算平均值"],
    ["work", "3", "multiple", "以下哪些是数据库的范式？", '{"A":"第一范式","B":"第二范式","C":"第三范式","D":"BC范式"}', "ABCD", "这些都是数据库的范式"],
    
    # 工作级 - 设计模式 (10道)
    ["work", "4", "single", "以下哪个设计模式用于处理对象间的一对多依赖关系？", '{"A":"单例模式","B":"观察者模式","C":"工厂模式","D":"装饰器模式"}', "B", "观察者模式用于处理对象间的一对多依赖关系，当一个对象状态改变时，所有依赖它的对象都会得到通知"],
    ["work", "4", "judge", "装饰器模式允许在不修改原有对象结构的情况下，动态地给对象添加额外的责任。", '{"A":"正确","B":"错误"}', "A", "装饰器模式的核心思想就是在不修改原有对象结构的情况下，动态地给对象添加额外的责任"],
    ["work", "4", "single", "以下哪个设计模式用于定义一系列算法，把它们封装起来，并使它们可以互相替换？", '{"A":"策略模式","B":"模板方法模式","C":"状态模式","D":"命令模式"}', "A", "策略模式用于定义一系列算法，把它们封装起来，并使它们可以互相替换"],
    ["work", "4", "multiple", "以下哪些是结构型设计模式？", '{"A":"适配器模式","B":"装饰器模式","C":"组合模式","D":"外观模式"}', "ABCD", "这些都属于结构型设计模式"],
    ["work", "4", "single", "以下哪个设计模式用于将一个复杂对象的构建与它的表示分离？", '{"A":"工厂模式","B":"建造者模式","C":"原型模式","D":"单例模式"}', "B", "建造者模式用于将一个复杂对象的构建与它的表示分离"],
    ["work", "4", "judge", "工厂方法模式定义了一个创建对象的接口，但由子类决定要实例化的类是哪一个。", '{"A":"正确","B":"错误"}', "A", "工厂方法模式的定义"],
    ["work", "4", "single", "以下哪个设计模式用于为一组对象提供一个一致的接口？", '{"A":"外观模式","B":"适配器模式","C":"桥接模式","D":"组合模式"}', "A", "外观模式用于为一组对象提供一个一致的接口"],
    ["work", "4", "multiple", "以下哪些是行为型设计模式？", '{"A":"命令模式","B":"迭代器模式","C":"观察者模式","D":"策略模式"}', "ABCD", "这些都属于行为型设计模式"],
    ["work", "4", "single", "以下哪个设计模式用于在不破坏封装性的前提下，捕获一个对象的内部状态，并在该对象之外保存这个状态？", '{"A":"命令模式","B":"备忘录模式","C":"状态模式","D":"中介者模式"}', "B", "备忘录模式用于捕获一个对象的内部状态，并在该对象之外保存这个状态"],
    ["work", "4", "judge", "抽象工厂模式提供一个创建一系列相关或相互依赖对象的接口，而无需指定它们具体的类。", '{"A":"正确","B":"错误"}', "A", "抽象工厂模式的定义"],
    
    # 专业级 - 算法基础 (5道)
    ["pro", "1", "single", "以下哪个算法用于解决0-1背包问题？", '{"A":"贪心算法","B":"动态规划","C":"回溯算法","D":"分支限界算法"}', "B", "0-1背包问题通常使用动态规划算法解决，贪心算法不能得到最优解"],
    ["pro", "1", "multiple", "以下哪些是NP完全问题？", '{"A":"旅行商问题","B":"背包问题","C":"排序问题","D":"图的着色问题"}', "ABD", "旅行商问题、背包问题、图的着色问题都是NP完全问题，排序问题是P问题"],
    ["pro", "1", "single", "以下哪个算法用于求解最长公共子序列问题？", '{"A":"贪心算法","B":"动态规划","C":"分治算法","D":"回溯算法"}', "B", "最长公共子序列问题通常使用动态规划算法解决"],
    ["pro", "1", "judge", "NP完全问题是NP问题中最难的一类。", '{"A":"正确","B":"错误"}', "A", "NP完全问题是NP问题中最难的一类"],
    ["pro", "1", "single", "以下哪个数据结构用于实现优先队列？", '{"A":"栈","B":"队列","C":"堆","D":"链表"}', "C", "堆用于实现优先队列"],
    
    # 专业级 - 编程语言 (5道)
    ["pro", "2", "single", "在C++中，以下哪个关键字用于定义常量？", '{"A":"const","B":"static","C":"final","D":"immutable"}', "A", "在C++中，使用const关键字定义常量"],
    ["pro", "2", "judge", "C++中的虚函数是通过虚函数表实现的。", '{"A":"正确","B":"错误"}', "A", "C++中的虚函数是通过虚函数表(vtable)实现的，每个包含虚函数的类都有一个虚函数表"],
    ["pro", "2", "single", "C++中，以下哪个关键字用于实现多态？", '{"A":"virtual","B":"override","C":"final","D":"abstract"}', "A", "C++中使用virtual关键字实现多态"],
    ["pro", "2", "multiple", "以下哪些是C++中的智能指针？", '{"A":"unique_ptr","B":"shared_ptr","C":"weak_ptr","D":"auto_ptr"}', "ABC", "unique_ptr、shared_ptr、weak_ptr都是C++中的智能指针"],
    ["pro", "2", "single", "C++中，以下哪个运算符用于动态内存分配？", '{"A":"malloc()","B":"new","C":"calloc()","D":"realloc()"}', "B", "C++中使用new运算符进行动态内存分配"],
    
    # 专业级 - 标准规范 (5道)
    ["pro", "3", "single", "以下哪个是NoSQL数据库？", '{"A":"MySQL","B":"PostgreSQL","C":"MongoDB","D":"Oracle"}', "C", "MongoDB是NoSQL数据库，其他选项都是关系型数据库"],
    ["pro", "3", "multiple", "以下哪些是分布式系统的特性？", '{"A":"一致性","B":"可用性","C":"分区容错性","D":"原子性"}', "ABC", "分布式系统的CAP理论包括一致性(Consistency)、可用性(Availability)、分区容错性(Partition tolerance)"],
    ["pro", "3", "single", "以下哪个是CAP理论中的权衡？", '{"A":"一致性和可用性","B":"可用性和分区容错性","C":"一致性和分区容错性","D":"以上都是"}', "D", "CAP理论中，只能同时满足其中两个特性"],
    ["pro", "3", "judge", "NoSQL数据库通常比关系型数据库更适合处理结构化数据。", '{"A":"正确","B":"错误"}', "B", "关系型数据库更适合处理结构化数据，NoSQL数据库更适合处理非结构化数据"],
    ["pro", "3", "single", "以下哪个是分布式共识算法？", '{"A":"Paxos","B":"Raft","C":"Gossip","D":"以上都是"}', "D", "这些都是分布式共识算法"],
    
    # 专业级 - 设计模式 (5道)
    ["pro", "4", "single", "以下哪个架构模式是基于服务的架构？", '{"A":"MVC","B":"微服务","C":"MVP","D":"MVVM"}', "B", "微服务是基于服务的架构，将应用拆分为多个独立的服务"],
    ["pro", "4", "subjective", "请简述RESTful API的设计原则。", "", "1. 资源标识：使用URI标识资源\n2. 统一接口：使用标准的HTTP方法\n3. 无状态：服务器不保存客户端状态\n4. 缓存：支持缓存以提高性能\n5. 分层系统：支持分层架构\n6. 按需编码：允许客户端下载并执行服务器代码（可选）", "RESTful API的设计原则包括资源标识、统一接口、无状态、缓存、分层系统和按需编码等"],
    ["pro", "4", "single", "以下哪个设计模式用于在运行时动态地给一个对象添加额外的责任？", '{"A":"装饰器模式","B":"代理模式","C":"适配器模式","D":"桥接模式"}', "A", "装饰器模式用于在运行时动态地给一个对象添加额外的责任"],
    ["pro", "4", "multiple", "以下哪些是微服务架构的特点？", '{"A":"服务独立部署","B":"服务独立扩展","C":"服务独立开发","D":"服务间通过API通信"}', "ABCD", "这些都是微服务架构的特点"],
    ["pro", "4", "single", "以下哪个设计模式用于将请求封装为对象，从而使您可以用不同的请求参数化客户端？", '{"A":"命令模式","B":"策略模式","C":"模板方法模式","D":"观察者模式"}', "A", "命令模式用于将请求封装为对象"],
]

# 写入题目数据
for row, question in enumerate(questions, 2):
    for col, value in enumerate(question, 1):
        worksheet[get_column_letter(col) + str(row)] = value

# 调整列宽
for col in range(1, len(headers) + 1):
    worksheet.column_dimensions[get_column_letter(col)].auto_size = True

# 保存工作簿
excel_path = "/root/CodeDojo/docs/questions_template.xlsx"
workbook.save(excel_path)
print(f"示例题库Excel文件已生成：{excel_path}")
print(f"文件包含 {len(questions)} 道题目")
